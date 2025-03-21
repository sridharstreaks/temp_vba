from openpyxl import load_workbook, Workbook

def get_wildcard(text):
    """
    Returns the text before the last '-' in the string.
    If no '-' is found, returns the full text.
    """
    if text is None:
        return ""
    return text.rsplit('-', 1)[0] if '-' in text else text

def read_column(ws, col_letter="A"):
    """
    Reads all cell values from the specified column in the worksheet.
    """
    return [cell.value for cell in ws[col_letter] if cell.value is not None]

def group_values(values):
    """
    Groups consecutive values where each new value contains the wildcard from the first element
    of the group. Only groups with 2 or more elements are returned.
    """
    groups = []
    current_group = []
    current_wildcard = None

    for value in values:
        # If starting a new group, calculate the wildcard from the current cell.
        if not current_group:
            current_group.append(value)
            current_wildcard = get_wildcard(value)
        else:
            # If current value contains the wildcard from the first element, add it to the group.
            if current_wildcard in value:
                current_group.append(value)
            else:
                # Only groups with 2 or more cells are valid.
                if len(current_group) >= 2:
                    groups.append(current_group)
                # Reset the group with the current cell.
                current_group = [value]
                current_wildcard = get_wildcard(value)
    # Check at the end of the loop.
    if len(current_group) >= 2:
        groups.append(current_group)
    return groups

def create_pairs(group):
    """
    Given a list of values in a group, creates a list of tuples representing all
    possible ordered pairs (crossmatch) within that group.
    For example, for group = [A, B, C]:
      returns [(A,B), (A,C), (B,A), (B,C), (C,A), (C,B)]
    """
    pairs = []
    n = len(group)
    for i in range(n):
        for j in range(n):
            if i != j:
                pairs.append((group[i], group[j]))
    return pairs

def write_pairs_to_sheet(wb, pairs, sheet_name="Output"):
    """
    Writes the list of pairs to a new sheet in the workbook.
    Each pair is written in two columns.
    """
    new_sheet = wb.create_sheet(sheet_name)
    new_sheet.append(["Value 1", "Value 2"])
    for pair in pairs:
        new_sheet.append(list(pair))

def main():
    # Specify your input file name
    input_file = "input.xlsx"
    
    # Load workbook and select the active sheet (adjust if needed)
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Read column A (change letter if needed)
    values = read_column(ws, col_letter="A")
    print("Read values:", values)
    
    # Group values based on wildcard logic
    groups = group_values(values)
    print("Found groups:", groups)
    
    # Create all pairwise combinations from groups
    all_pairs = []
    for group in groups:
        group_pairs = create_pairs(group)
        all_pairs.extend(group_pairs)
    
    print("Total pairs created:", len(all_pairs))
    
    # Write the pairs to a new sheet in the workbook
    write_pairs_to_sheet(wb, all_pairs, sheet_name="Output")
    
    # Save the workbook (change the output file name if desired)
    output_file = "output.xlsx"
    wb.save(output_file)
    print(f"Output written to {output_file}")

if __name__ == "__main__":
    main()
