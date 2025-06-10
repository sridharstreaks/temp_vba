from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def insert_and_fill(
    path: str,
    sheet_name: str,
    start_row: int,
    interval: int,
    lookup_sheet_name: str = None
):
    # 1) Load workbook
    wb = load_workbook(path)
    ws = wb[sheet_name]
    
    # 2) Determine lookup sheet
    if lookup_sheet_name:
        ws_lookup = wb[lookup_sheet_name]
    else:
        idx = wb.sheetnames.index(sheet_name)
        ws_lookup = wb[wb.sheetnames[idx + 1]]
    
    # 3) Find last used row by looking at column A
    last_row = max(
        cell.row for cell in ws['A'] 
        if cell.value is not None
    )
    # 4) Find last header column in row 1
    last_col = max(cell.column for cell in ws[1] if cell.value is not None)
    
    # 5) Compute insertion points
    max_k = (last_row - start_row) // interval
    insert_points = [
        start_row + k * interval
        for k in range(max_k + 1)
        if start_row + k * interval <= last_row
    ]
    # reverse so earlier inserts don't shift the later ones
    for r in reversed(insert_points):
        # insert blank row
        ws.insert_rows(r)
        
        # copy A:D from row r+1 → row r
        for col in range(1, 5):  # 1=A ... 4=D
            ws.cell(r, col).value = ws.cell(r + 1, col).value
        
        # set E to "blah"
        ws.cell(r, 5).value = "blah"
        
        # leave F-H alone
        
        # paste XLOOKUP formula in cols I (=9) → last_col
        for col in range(9, last_col + 1):
            header_addr = f"{get_column_letter(col)}$1"
            # formula references: C<r> and header in row 1
            c_ref       = f"C{r}"
            lookup_name = ws_lookup.title
            # wrap sheet name in single-quotes if needed
            if ' ' in lookup_name:
                lookup_name = f"'{lookup_name}'"
            # build formula
            formula = (
                f"=XLOOKUP(1,"
                f"({lookup_name}!C:C={c_ref})*"
                f"({lookup_name}!A:A={header_addr}),"
                f"{lookup_name}!E:E,"
                f"\"\",0)"
            )
            ws.cell(r, col).value = formula

    # 6) Save
    wb.save(path)


if __name__ == "__main__":
    insert_and_fill(
        path="C:/path/to/your_workbook.xlsx",
        sheet_name="Sheet1",
        start_row=2,
        interval=3,
        # optionally: lookup_sheet_name="LookupData"
    )
