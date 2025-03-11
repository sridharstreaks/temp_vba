import pandas as pd
from openpyxl import load_workbook
import requests
import json

# ----- Step 1: Read the Excel file and extract ISBNs from column A -----
excel_file = input("Path of the excel file: ").replace("\"", "")
df = pd.read_excel(excel_file)
all_isbns = df['isbn'].tolist()  # Get the ISBNs (adjust column name as needed)

# ----- Step 2: Ask for start and end row and create the selected ISBNs list -----
start_row = int(input("Enter start row number: "))
end_row = int(input("Enter end row number: "))

# Adjust for header row if necessary.
# In this example, if start_row==1 we assume no header; otherwise adjust for header.
if start_row == 1:
    isbns = all_isbns[start_row - 1: end_row]
else:
    isbns = all_isbns[start_row - 2: end_row - 1]

# ----- Step 3: Process ISBNs in batches and extract data from Kennys site -----
batch_size = 5  # You can change the batch size if needed
extracted_data = {}

# Load the workbook once before processing batches.
wb = load_workbook(excel_file)
ws = wb.active  # Assumes the data is in the active sheet

# Process ISBNs in batches
for i in range(0, len(isbns), batch_size):
    # Define current batch
    batch = isbns[i : i + batch_size]
    
    # Process each ISBN in the current batch
    for isbn in batch:
        response = requests.get(f"https://www.kennys.ie/searchquery?size=12&from=0&source=%7B%22query%22%3A%7B%22bool%22%3A%7B%22must%22%3A%5B%7B%22terms%22%3A%7B%22state%22%3A%5B1%2C2%5D%7D%7D%2C%7B%22terms%22%3A%7B%22cat_state%22%3A%5B1%2C2%5D%7D%7D%2C%7B%22bool%22%3A%7B%22should%22%3A%5B%7B%22match%22%3A%7B%22title%22%3A%7B%22query%22%3A%22{isbn}%22%2C%22operator%22%3A%22and%22%2C%22boost%22%3A1.7%7D%7D%7D%2C%7B%22match%22%3A%7B%22description%22%3A%7B%22query%22%3A%22{isbn}%22%2C%22operator%22%3A%22and%22%2C%22boost%22%3A0.7%7D%7D%7D%2C%7B%22match%22%3A%7B%22path%22%3A%7B%22query%22%3A%22{isbn}%22%2C%22operator%22%3A%22and%22%2C%22boost%22%3A2%7D%7D%7D%5D%7D%7D%2C%7B%22bool%22%3A%7B%22should%22%3A%5B%7B%22range%22%3A%7B%22end_date%22%3A%7B%22gte%22%3A%222025-03-09+11%3A27%3A54%22%7D%7D%7D%2C%7B%22term%22%3A%7B%22end_date%22%3A%220%22%7D%7D%2C%7B%22bool%22%3A%7B%22must_not%22%3A%5B%7B%22exists%22%3A%7B%22field%22%3A%22end_date%22%7D%7D%5D%7D%7D%5D%7D%7D%2C%7B%22bool%22%3A%7B%22should%22%3A%5B%7B%22range%22%3A%7B%22publish_end_date%22%3A%7B%22gte%22%3A%222025-03-09+11%3A27%3A54%22%7D%7D%7D%2C%7B%22term%22%3A%7B%22publish_end_date%22%3A%220%22%7D%7D%2C%7B%22bool%22%3A%7B%22must_not%22%3A%5B%7B%22exists%22%3A%7B%22field%22%3A%22publish_end_date%22%7D%7D%5D%7D%7D%5D%7D%7D%2C%7B%22bool%22%3A%7B%22should%22%3A%5B%7B%22range%22%3A%7B%22publish_start_date%22%3A%7B%22lte%22%3A%222025-03-09+11%3A27%3A54%22%7D%7D%7D%2C%7B%22term%22%3A%7B%22publish_start_date%22%3A%220%22%7D%7D%2C%7B%22bool%22%3A%7B%22must_not%22%3A%5B%7B%22exists%22%3A%7B%22field%22%3A%22publish_start_date%22%7D%7D%5D%7D%7D%5D%7D%7D%5D%2C%22should%22%3A%5B%5D%2C%22must_not%22%3A%5B%5D%2C%22filter%22%3A%5B%5D%7D%7D%2C%22suggest%22%3A%7B%22kennysdidyoumean%22%3A%7B%22text%22%3A%22{isbn}%22%2C%22phrase%22%3A%7B%22field%22%3A%22description%22%7D%7D%7D%7D&source_content_type=application%2Fjson&type=")
        
        if response.status_code == 200:
            data = response.content
        else:
            extracted_data[isbn] = "Request failed"
            print(isbn, "Request failed")
            continue

        # Parse JSON data if needed.
        if isinstance(data, bytes):
            data = json.loads(data)

        # Check for a successful shard and match ISBN from the suggest block.
        if data.get("_shards", {}).get("successful") != 0:
            #print(data)
            isbn_returned = data.get("suggest", {}).get("kennysdidyoumean", [])[0].get("text")
            if isbn_returned == isbn:
                hits_list = data.get("hits", {}).get("hits", [])
                if not hits_list:
                    extracted_result = "No results found"
                else:
                    first_hit = hits_list[0]
                    source = first_hit.get("_source", {})
                    path_value = source.get("path")
                    extracted_result = "https://www.kennys.ie/" + path_value
            else:
                extracted_result = "isbn not match"
        else:
            extracted_result = "Failed to retrieve data"

        extracted_data[isbn] = extracted_result
        print(isbn, extracted_result)
    
    # After processing the current batch, write its results to Excel.
    for j, isbn in enumerate(batch):
        # Calculate the Excel row corresponding to this ISBN.
        excel_row = start_row + i + j
        ws.cell(row=excel_row, column=3, value=extracted_data.get(isbn, ""))
    
    # Save the workbook after each batch.
    wb.save(excel_file)
    print(f"Batch from row {start_row + i} to row {start_row + i + len(batch) - 1} has been saved.")

print("All results have been processed and saved to Excel.")
