import os
import sys
import glob
to_datetime = None
from datetime import datetime, timedelta
import openpyxl

def get_recent_files(folder, minutes=5):
    now = datetime.now()
    pattern = os.path.join(folder, '*.xlsx')
    files = glob.glob(pattern)
    recent = []
    for f in files:
        modified = datetime.fromtimestamp(os.path.getmtime(f))
        if now - modified <= timedelta(minutes=minutes):
            recent.append(f)
    return recent

def process_files(folder):
    files = get_recent_files(folder)
    if len(files) != 2:
        raise Exception(f"Expected exactly 2 files updated in the last 5 minutes, found {len(files)}.")
    # Sort for consistency
    files.sort()
    first_file = files[0]
    wb = openpyxl.load_workbook(first_file)
    ws = wb.active
    # Insert new column A
    ws.insert_cols(1)
    # Set header
    ws['A1'] = 'Segment'
    # Determine the last row with data in existing columns (after insert -> original B...)
    max_row = ws.max_row
    for row in range(2, max_row + 1):
        # Only fill where other columns have data
        if any(ws.cell(row=row, column=col).value is not None for col in range(2, ws.max_column + 1)):
            ws.cell(row=row, column=1).value = 'yes'
    # Save and close
    wb.save(first_file)

if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage: python script.py <folder_path>")
        sys.exit(1)
    folder = sys.argv[1]
    process_files(folder)
    print("Processing completed.")
