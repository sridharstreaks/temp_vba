# import sys
# import time
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import ChromiumOptions
# from webdriver_manager.chrome import ChromeDriverManager

# chrome_options = ChromiumOptions()

# service = Service(ChromeDriverManager().install())

# driver = webdriver.Chrome(chrome_options=chrome_options, service=service)
# driver.get("http://www.python.org")

# time.sleep(sys.maxsize)
# driver.quit()

from helium import *
import time
import pandas as pd
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

# ----- Step 1: Read the Excel file and extract URLs from column A -----
excel_file = input("Path of the excel file: ")  # Path to your Excel file
df = pd.read_excel(excel_file)
urls = df.iloc[:, 0].tolist()  # Get all URLs from the first column

# ----- Step 2: Ask for start and end row and create the selected URLs list -----
# Note: Excel rows are 1-indexed so we adjust for Python's 0-indexing.
start_row = int(input("Enter start row number: "))
end_row = int(input("Enter end row number: "))
urls = urls[start_row - 1 : end_row]  # Slice the list based on user input

# XPath of the element to extract text and href from
xpath = "//h1[@class='result-title']//a"

# Dictionary to store extracted text and href
extracted_data = {}

driver = get_driver()  # Helium's driver instance

# Wait for element text to change
element = driver.find_element(By.ID, "element-id")
initial_text = element.text

# Start the browser
start_chrome(headless=False)

for url in urls:
    # Navigate to the URL
    go_to(url)

    # Wait for a specified time (e.g., 5 seconds)
    time.sleep(8)

    
    # Find the element using XPath
    element = S(xpath)
    
    # Extract text and href if the element exists
    if element.exists():
        text = element.web_element.text
        href = element.web_element.get_attribute('href')
        extracted_data[text] = href
    else:
        extracted_data[url] = "No results found"

# Close the browser
kill_browser()

# ----- Step 4: Wait for user confirmation before writing back to the Excel file -----
confirmation = input("Do you want to write the results back to Excel? (yes/no): ").strip().lower()
if confirmation == "yes":
    # Load the workbook with openpyxl
    wb = load_workbook(excel_file)
    ws = wb.active  # Assumes the data is in the active sheet

    # Write each result to column B (adjust the column if needed) starting at the specified start_row.
    current_row = start_row
    for url in urls:
        ws.cell(row=current_row, column=2, value=extracted_data.get(url, ""))
        current_row += 1

    wb.save(excel_file)
    print("Results have been written to Excel.")
else:
    print("Operation cancelled. Results not written to Excel.")